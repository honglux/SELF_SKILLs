import os
import re
import sys
import csv


def extract_cases(filepath, start_line, end_line):
    # 智能读取：先尝试 UTF-8，如果碰到不兼容的（比如旧版 C++ 里的中文注释），自动切换到 GBK
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            all_lines = f.readlines()
    except UnicodeDecodeError:
        with open(filepath, "r", encoding="gbk", errors="ignore") as f:
            all_lines = f.readlines()

    total_lines = len(all_lines)
    if start_line < 1 or end_line > total_lines or start_line > end_line:
        print(f"错误：行号范围无效。文件共 {total_lines} 行，输入范围 {start_line}-{end_line}。")
        sys.exit(1)

    target_lines = all_lines[start_line - 1 : end_line]

    case_re = re.compile(r"^\s*case\s+(.+?)\s*:")
    default_re = re.compile(r"^\s*default\s*:")

    # 计算目标范围之前的括号深度
    base_depth = 0
    for line in all_lines[: start_line - 1]:
        base_depth += line.count("{") - line.count("}")

    # 计算目标范围内每一行开始时的括号深度
    depth_at_start = [base_depth]
    for line in target_lines:
        depth_at_start.append(depth_at_start[-1] + line.count("{") - line.count("}"))
    # depth_at_start[i] = 处理 target_lines[i] 之前的深度

    # 找出范围内所有 case 行的索引
    case_indices = [i for i, line in enumerate(target_lines) if case_re.match(line)]

    cases = []
    for pos, idx in enumerate(case_indices):
        start_line_num = start_line + idx
        case_depth = depth_at_start[idx]

        # 提取 case 标签
        m = case_re.match(target_lines[idx])
        case_label = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', m.group(1).strip())
        if case_label.startswith(('=', '+', '-', '@')):
            case_label = f"'{case_label}"

        # 找结束行：下一个同级 case/default 的上一行，或深度下降时所在的 }
        end_idx = len(target_lines) - 1
        for j in range(idx + 1, len(target_lines)):
            if depth_at_start[j] == case_depth:
                if case_re.match(target_lines[j]) or default_re.match(target_lines[j]):
                    end_idx = j - 1
                    break
            if depth_at_start[j + 1] < case_depth:
                end_idx = j
                break

        end_line_num = start_line + end_idx
        cases.append((start_line_num, end_line_num, case_label))

    return cases


def write_csv(file_abs_path, cases, output_path):
    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["绝对路径", "起始行号", "结束行号", "case标签"])
        for start_line_num, end_line_num, label in cases:
            writer.writerow([file_abs_path, start_line_num, end_line_num, label])


def write_xlsx(file_abs_path, cases, output_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Statements"

    ws["A1"] = "绝对路径"
    ws["B1"] = "起始行号"
    ws["C1"] = "结束行号"
    ws["D1"] = "case标签"

    for row_idx, (start_line_num, end_line_num, label) in enumerate(cases, start=2):
        ws.cell(row=row_idx, column=1, value=file_abs_path)
        ws.cell(row=row_idx, column=2, value=start_line_num)
        ws.cell(row=row_idx, column=3, value=end_line_num)
        ws.cell(row=row_idx, column=4, value=label)

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    wb.save(output_path)


def main():
    if len(sys.argv) < 4 or len(sys.argv) > 5:
        print("用法: python extract_cases.py <文件路径> <起始行号> <结束行号> [csv|xlsx]")
        print("  csv  - 输出 CSV 文件（默认，无需 openpyxl）")
        print("  xlsx - 输出 Excel 文件（需安装 openpyxl）")
        sys.exit(1)

    filepath = sys.argv[1]
    try:
        start_line = int(sys.argv[2])
        end_line = int(sys.argv[3])
    except ValueError:
        print("错误：行号必须为整数。")
        sys.exit(1)

    fmt = sys.argv[4].lower() if len(sys.argv) == 5 else "csv"
    if fmt not in ("csv", "xlsx"):
        print(f"错误：不支持的格式 '{fmt}'，仅支持 csv 或 xlsx。")
        sys.exit(1)

    print("正在扫描...")
    cases = extract_cases(filepath, start_line, end_line)

    file_abs_path = os.path.abspath(filepath)
    output_path = f"case_statements.{fmt}"

    if fmt == "csv":
        write_csv(file_abs_path, cases, output_path)
    else:
        write_xlsx(file_abs_path, cases, output_path)

    print(f"提取到 {len(cases)} 个 case 标签，结果已保存至 {output_path}")


if __name__ == "__main__":
    main()
