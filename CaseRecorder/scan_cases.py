import argparse
import csv
import logging
import os
import sys
import time
from pathlib import Path

from ai_client import ClaudeCodeClient, AICallError

logger = logging.getLogger("scan_cases")


def setup_logging():
    logger.setLevel(logging.DEBUG)

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)-5s %(message)s", datefmt="%H:%M:%S"))
    logger.addHandler(ch)

    fh = logging.FileHandler("scan.log", encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(logging.Formatter("[%(asctime)s] %(levelname)-5s %(message)s"))
    logger.addHandler(fh)


def load_prompt(path):
    with open(path, "r", encoding="utf-8") as f:
        text = f.read()
    return "".join(ch for ch in text if ch.isprintable() or ch in "\n\r\t")


def read_input(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _read_csv(path)
    elif ext == ".xlsx":
        return _read_xlsx(path)
    else:
        print(f"错误：不支持的输入格式 '{ext}'，仅支持 .csv 或 .xlsx。")
        sys.exit(1)


def _read_csv(path):
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if header is None:
            print("错误：输入文件为空。")
            sys.exit(1)
        rows = []
        for row in reader:
            if len(row) >= 4:
                rows.append((row[0], row[1], row[2], row[3]))
    return rows


def _read_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None and row[2] is not None and row[3] is not None:
            rows.append((str(row[0]), str(row[1]), str(row[2]), str(row[3])))
    return rows


class _CsvWriter:
    """增量写入 CSV：每行结果立即落盘并 flush，中途崩溃不丢已完成的记录。"""

    def __init__(self, path):
        self._f = open(path, "w", encoding="utf-8-sig", newline="")
        self._w = csv.writer(self._f)
        self._w.writerow(["绝对路径", "起始行号", "case标签", "AI返回"])
        self._f.flush()

    def add_row(self, row):
        self._w.writerow(row)
        self._f.flush()

    def close(self):
        self._f.close()


class _XlsxWriter:
    """批量写入 xlsx：openpyxl 不支持增量追加，结束时一次性保存。"""

    def __init__(self, path):
        self._path = path
        self._rows = []

    def add_row(self, row):
        self._rows.append(row)

    def close(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Scan Results"
        ws["A1"] = "绝对路径"
        ws["B1"] = "起始行号"
        ws["C1"] = "case标签"
        ws["D1"] = "AI返回"
        for i, row in enumerate(self._rows, start=2):
            ws.cell(row=i, column=1, value=row[0])
            ws.cell(row=i, column=2, value=row[1])
            ws.cell(row=i, column=3, value=row[2])
            ws.cell(row=i, column=4, value=row[3])
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 80
        wb.save(self._path)


def open_output(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        return _CsvWriter(path)
    elif ext == ".xlsx":
        logger.info("xlsx 模式将在内存中累积结果，大数据量（>500 行）建议使用 CSV")
        return _XlsxWriter(path)
    else:
        print(f"错误：不支持的输出格式 '{ext}'，仅支持 .csv 或 .xlsx。")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="AI 逐 Case 扫描器")
    parser.add_argument("--input", required=True, help="extract_cases.py 生成的 CSV 或 xlsx 文件路径")
    parser.add_argument("--prompt-template", required=True, help="Prompt 模板文件路径（Markdown）")
    parser.add_argument("--output", default=None, help="输出文件路径（默认与输入同格式，文件名为 scan_results）")
    parser.add_argument("--auth-hints", default="无", help="用户指定的鉴权函数名列表（可选，多个用逗号分隔）")
    args = parser.parse_args()

    input_path = args.input
    if not os.path.isfile(input_path):
        print(f"错误：输入文件不存在: {input_path}")
        sys.exit(1)

    prompt_path = args.prompt_template
    if not os.path.isfile(prompt_path):
        print(f"错误：Prompt 模板文件不存在: {prompt_path}")
        sys.exit(1)

    setup_logging()

    logger.info("scan_cases 启动")
    logger.info("输入文件: %s", os.path.abspath(input_path))
    logger.info("Prompt 模板: %s", os.path.abspath(prompt_path))
    logger.info("鉴权函数提示: %s", args.auth_hints)

    prompt_template = load_prompt(prompt_path)
    logger.debug("Prompt 模板长度: %d 字符", len(prompt_template))

    input_rows = read_input(input_path)
    logger.info("读取到 %d 行数据", len(input_rows))

    if args.output is None:
        ext = os.path.splitext(input_path)[1]
        args.output = f"scan_results{ext}"

    client = ClaudeCodeClient()
    writer = open_output(args.output)

    total = len(input_rows)
    failed = 0

    try:
        for i, (file_path, start_line, end_line, case_label) in enumerate(input_rows, start=1):
            logger.info("[%d/%d] 扫描: %s 行 %s-%s case %s", i, total, file_path, start_line, end_line, case_label)
            t0 = time.time()

            prompt = (prompt_template
                      .replace("{file_path}", file_path)
                      .replace("{start_line}", start_line)
                      .replace("{end_line}", end_line)
                      .replace("{case_label}", case_label)
                      .replace("{auth_hints}", args.auth_hints))

            workdir = Path(os.path.dirname(file_path))
            if not workdir.is_dir():
                workdir = Path.cwd()
                logger.debug("源文件目录不存在，使用当前目录作为 workdir: %s", workdir)

            try:
                response = client.invoke(prompt, workdir)
                writer.add_row((file_path, start_line, case_label, response))
                elapsed = time.time() - t0
                logger.info("[%d/%d] 完成 (耗时 %.0fs)", i, total, elapsed)
            except Exception as e:
                logger.error("[%d/%d] AI 调用失败: %s", i, total, e)
                writer.add_row((file_path, start_line, case_label, f"[AI调用失败] {e}"))
                failed += 1
    finally:
        writer.close()

    ok = total - failed
    logger.info("全部扫描完成: 成功 %d, 失败 %d, 共 %d 行, 结果保存在 %s", ok, failed, total, os.path.abspath(args.output))


if __name__ == "__main__":
    main()
