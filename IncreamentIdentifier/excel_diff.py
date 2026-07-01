"""Excel 去重对比工具。

对比两个文件夹中的 .xlsx 文件，对 folder_b 中的每一行标记
是否已在 folder_a 中出现过。
"""

import sys
from pathlib import Path

from openpyxl import load_workbook

HEADER_SCANED_OBJECT = "Scaned Object"
HEADER_MATCH_CONTENT = "Match Content"
HEADER_MATCH_CONTEXT = "Match Context"
SHEET_NAME = "Detail_1"
DUPLICATE_COLUMN = "Duplicate"


def extract_filename(path_str: str) -> str:
    """从文件路径（Linux 或 Windows）中提取文件名。"""
    if not path_str:
        return ""
    # 同时处理 / 和 \
    return path_str.replace("\\", "/").rstrip("/").split("/")[-1]


def build_keyset(folder: Path) -> set[tuple[str, str, str]]:
    """构建 folder_a 的 key 集合。每个 key 为 (文件名, MatchContent, MatchContext) 三元组。"""
    keys: set[tuple[str, str, str]] = set()
    xlsx_files = sorted(folder.glob("*.xlsx"))
    if not xlsx_files:
        print(f"警告: {folder} 中没有找到 .xlsx 文件")

    for xlsx_path in xlsx_files:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        if SHEET_NAME not in wb.sheetnames:
            print(f"错误: {xlsx_path.name} 中不存在 sheet \"{SHEET_NAME}\"")
            sys.exit(1)

        ws = wb[SHEET_NAME]

        # 查找三列的表头位置（第 1 行）
        col_idx_scan = None
        col_idx_content = None
        col_idx_context = None
        for cell in ws[1]:
            if cell.value == HEADER_SCANED_OBJECT:
                col_idx_scan = cell.column
            elif cell.value == HEADER_MATCH_CONTENT:
                col_idx_content = cell.column
            elif cell.value == HEADER_MATCH_CONTEXT:
                col_idx_context = cell.column

        if col_idx_scan is None or col_idx_content is None or col_idx_context is None:
            print(f"错误: {xlsx_path.name} 的 sheet \"{SHEET_NAME}\" 中缺少必要的表头列")
            sys.exit(1)

        # 读取数据行（从第 2 行开始），向前填充合并单元格的空值
        last_scan_obj = ""
        for row in ws.iter_rows(min_row=2, values_only=True):
            # openpyxl values_only 模式下，列索引从 0 开始
            scan_obj = row[col_idx_scan - 1]
            match_content = row[col_idx_content - 1]
            match_context = row[col_idx_context - 1]

            # forward fill
            if scan_obj is None:
                scan_obj = last_scan_obj
            else:
                scan_obj = str(scan_obj)
                last_scan_obj = scan_obj

            match_content = str(match_content) if match_content is not None else ""
            match_context = str(match_context) if match_context is not None else ""

            filename = extract_filename(scan_obj)
            keys.add((filename, match_content, match_context))

        wb.close()

    return keys


def process_folder_b(folder_b: Path, keyset: set[tuple[str, str, str]], output_dir: Path):
    """处理 folder_b 中的文件，追加 Duplicate 列后保存到 output_dir。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_files = sorted(folder_b.glob("*.xlsx"))
    if not xlsx_files:
        print(f"警告: {folder_b} 中没有找到 .xlsx 文件")
        return

    for xlsx_path in xlsx_files:
        wb = load_workbook(xlsx_path)
        if SHEET_NAME not in wb.sheetnames:
            print(f"错误: {xlsx_path.name} 中不存在 sheet \"{SHEET_NAME}\"")
            sys.exit(1)

        ws = wb[SHEET_NAME]

        # 查找三列表头位置
        col_idx_scan = None
        col_idx_content = None
        col_idx_context = None
        for cell in ws[1]:
            if cell.value == HEADER_SCANED_OBJECT:
                col_idx_scan = cell.column
            elif cell.value == HEADER_MATCH_CONTENT:
                col_idx_content = cell.column
            elif cell.value == HEADER_MATCH_CONTEXT:
                col_idx_context = cell.column

        if col_idx_scan is None or col_idx_content is None or col_idx_context is None:
            print(f"错误: {xlsx_path.name} 的 sheet \"{SHEET_NAME}\" 中缺少必要的表头列")
            sys.exit(1)

        # 找到最后一列（包含隐藏列），在右侧追加新列
        max_col = ws.max_column
        dup_col = max_col + 1

        # 写入表头
        ws.cell(row=1, column=dup_col, value=DUPLICATE_COLUMN)

        # 遍历数据行
        last_scan_obj = ""
        last_row = ws.max_row
        for row_idx in range(2, last_row + 1):
            scan_cell = ws.cell(row=row_idx, column=col_idx_scan)
            content_cell = ws.cell(row=row_idx, column=col_idx_content)
            context_cell = ws.cell(row=row_idx, column=col_idx_context)

            scan_obj = scan_cell.value
            if scan_obj is None:
                scan_obj = last_scan_obj
            else:
                scan_obj = str(scan_obj)
                last_scan_obj = scan_obj

            match_content = str(content_cell.value) if content_cell.value is not None else ""
            match_context = str(context_cell.value) if context_cell.value is not None else ""

            filename = extract_filename(scan_obj)
            key = (filename, match_content, match_context)
            ws.cell(row=row_idx, column=dup_col, value=key in keyset)

        output_path = output_dir / xlsx_path.name
        wb.save(output_path)
        print(f"已处理: {xlsx_path.name} -> {output_path}")
        wb.close()


def main():
    if len(sys.argv) != 3:
        print("用法: python excel_diff.py <folder_a> <folder_b>")
        print("示例: python excel_diff.py D:\\data\\baseline D:\\data\\newscan")
        sys.exit(1)

    folder_a = Path(sys.argv[1])
    folder_b = Path(sys.argv[2])

    if not folder_a.is_dir():
        print(f"错误: {folder_a} 不是有效的文件夹")
        sys.exit(1)
    if not folder_b.is_dir():
        print(f"错误: {folder_b} 不是有效的文件夹")
        sys.exit(1)

    output_dir = folder_b.parent / f"{folder_b.name}_marked"

    print(f"基准文件夹: {folder_a}")
    print(f"待处理文件夹: {folder_b}")
    print(f"输出文件夹: {output_dir}")
    print()

    print("正在构建基准 key 集合...")
    keyset = build_keyset(folder_a)
    print(f"已构建 {len(keyset)} 个 key")
    print()

    print("正在处理 folder_b 并标记重复项...")
    process_folder_b(folder_b, keyset, output_dir)
    print()
    print("完成。")


if __name__ == "__main__":
    main()
