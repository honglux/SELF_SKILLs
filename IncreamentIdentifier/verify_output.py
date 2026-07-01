"""验证 excel_diff.py 的输出是否正确"""
from pathlib import Path
from openpyxl import load_workbook


def verify():
    output_dir = Path("D:/scripts/test_data/newscan_marked")
    if not output_dir.is_dir():
        print("输出文件夹不存在!")
        return

    all_ok = True

    # --- 验证 file1.xlsx ---
    wb = load_workbook(output_dir / "file1.xlsx")
    ws = wb["Detail_1"]

    # 找 Duplicate 列
    dup_col = None
    for cell in ws[1]:
        if cell.value == "Duplicate":
            dup_col = cell.column
            break

    if dup_col is None:
        print("FAIL: file1.xlsx 缺少 Duplicate 列")
        all_ok = False
    else:
        # 读取每行的 Duplicate 值和关键信息
        results = {}
        for row_idx in range(2, ws.max_row + 1):
            b = ws.cell(row=row_idx, column=2).value
            r = ws.cell(row=row_idx, column=18).value
            s = ws.cell(row=row_idx, column=19).value
            dup = ws.cell(row=row_idx, column=dup_col).value
            results[row_idx] = (b, r, s, dup)

        # 预期:
        # row2: (C:\project\src\main.java, password=12345, line 10 context) -> True (完全匹配)
        # row3: (main.java ff, different content, different context) -> False (三项中两项不同)
        # row4: (main.java ff, SELECT * FROM users, line 999 different) -> False (Context 不同)
        # row5: (main.java ff, api_key="abcxyz", line 30 // 中文注释) -> True (完全匹配, 合并单元格 ff)
        # row6: (/new/file.cpp, new content, new context) -> False (全新行)

        if results[2][3] is not True:
            print(f"FAIL: file1.xlsx row2 期望 True, 实际 {results[2][3]}")
            all_ok = False
        else:
            print("PASS: file1.xlsx row2 -> True (完全匹配)")

        if results[3][3] is not False:
            print(f"FAIL: file1.xlsx row3 期望 False, 实际 {results[3][3]}")
            all_ok = False
        else:
            print("PASS: file1.xlsx row3 -> False (Match Content + Context 均不同)")

        if results[4][3] is not False:
            print(f"FAIL: file1.xlsx row4 期望 False, 实际 {results[4][3]}")
            all_ok = False
        else:
            print("PASS: file1.xlsx row4 -> False (Match Context 不同)")

        if results[5][3] is not True:
            print(f"FAIL: file1.xlsx row5 期望 True, 实际 {results[5][3]}")
            all_ok = False
        else:
            print("PASS: file1.xlsx row5 -> True (完全匹配, 合并单元格 forward fill)")

        if results[6][3] is not False:
            print(f"FAIL: file1.xlsx row6 期望 False, 实际 {results[6][3]}")
            all_ok = False
        else:
            print("PASS: file1.xlsx row6 -> False (全新行)")

        # 验证原始内容保留：B列合并单元格是否还在
        merged = ws.merged_cells.ranges
        if merged:
            print(f"PASS: file1.xlsx 保留了合并单元格: {[str(m) for m in merged]}")
        else:
            print("FAIL: file1.xlsx 合并单元格丢失")
            all_ok = False

        # 验证隐藏列/无关列保留：第3列 SomeColumn 是否还在
        if ws.cell(row=1, column=3).value == "SomeColumn":
            print("PASS: file1.xlsx 保留了额外列 (SomeColumn)")
        else:
            print("FAIL: file1.xlsx 额外列丢失")
            all_ok = False

    wb.close()

    # --- 验证 file3.xlsx ---
    wb = load_workbook(output_dir / "file3.xlsx")
    ws = wb["Detail_1"]

    dup_col = None
    for cell in ws[1]:
        if cell.value == "Duplicate":
            dup_col = cell.column
            break

    if dup_col is None:
        print("FAIL: file3.xlsx 缺少 Duplicate 列")
        all_ok = False
    else:
        results = {}
        for row_idx in range(2, ws.max_row + 1):
            b = ws.cell(row=row_idx, column=2).value
            r = ws.cell(row=row_idx, column=18).value
            s = ws.cell(row=row_idx, column=19).value
            dup = ws.cell(row=row_idx, column=dup_col).value
            results[row_idx] = (b, r, s, dup)

        # row2: (/opt/app/run.sh, #!/bin/bash, line 1 shebang) -> True (在 folder_a file2 中存在)
        if results[2][3] is not True:
            print(f"FAIL: file3.xlsx row2 期望 True, 实际 {results[2][3]}")
            all_ok = False
        else:
            print("PASS: file3.xlsx row2 -> True (folder_a file2 中存在)")

        # row3: (/opt/app/run.sh, echo hello, line 10) -> False
        if results[3][3] is not False:
            print(f"FAIL: file3.xlsx row3 期望 False, 实际 {results[3][3]}")
            all_ok = False
        else:
            print("PASS: file3.xlsx row3 -> False (Match Content 不同)")

        # row4: (/unknown/file.txt, some content, some context) -> False
        if results[4][3] is not False:
            print(f"FAIL: file3.xlsx row4 期望 False, 实际 {results[4][3]}")
            all_ok = False
        else:
            print("PASS: file3.xlsx row4 -> False (全新行)")

    wb.close()

    # --- 验证 key 数量 ---
    keyset_file = output_dir / "file1.xlsx"
    keyset_file3 = output_dir / "file3.xlsx"
    print(f"\n输出文件: {keyset_file.name}, {keyset_file3.name}")

    if all_ok:
        print("\n全部 PASS!")
    else:
        print("\n存在 FAIL!")


if __name__ == "__main__":
    verify()
