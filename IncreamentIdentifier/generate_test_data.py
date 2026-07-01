"""生成测试数据"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def make_test_data():
    base = Path("D:/scripts/test_data")
    folder_a = base / "baseline"
    folder_b = base / "newscan"
    for d in [folder_a, folder_b]:
        d.mkdir(parents=True, exist_ok=True)

    # ========== folder_a: file1.xlsx ==========
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail_1"

    # 写表头: A=ignored, B=Scaned Object, ..., R=Match Content, S=Match Context
    # 中间加一些无关列模拟真实场景
    headers = {
        1: "ID",
        2: "Scaned Object",
        3: "SomeColumn",
        18: "Match Content",
        19: "Match Context",
    }
    for col, val in headers.items():
        ws.cell(row=1, column=col, value=val)

    # 数据行：含合并单元格（B列）, Windows/Linux 路径, 中文, 特殊字符
    rows_a1 = [
        # row2: B="C:\project\src\main.java", R="password=12345", S="line 10 context"
        {2: r"C:\project\src\main.java", 3: "x", 18: "password=12345", 19: "line 10 context"},
        # row3: B 合并(空), 同一个文件的不同匹配
        {2: None, 3: "x", 18: r"SELECT * FROM users", 19: "line 20 context"},
        # row4: B 合并(空)
        {2: None, 3: "x", 18: "api_key = \"abcxyz\"", 19: "line 30 // 中文注释"},
        # row5: B 新文件 Linux 路径
        {2: "/home/user/config.yaml", 3: "x", 18: "db_password: secret!!", 19: "包含中文和特殊字符\t\n"},
        # row6: B 合并(空)
        {2: None, 3: "x", 18: "TODO: fix me", 19: "context with $pecial ch@rs"},
    ]
    for i, row_data in enumerate(rows_a1, start=2):
        for col, val in row_data.items():
            if val is not None:
                ws.cell(row=i, column=col, value=val)

    # 手动合并 B 列单元格
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)
    ws.merge_cells(start_row=5, start_column=2, end_row=6, end_column=2)

    wb.save(folder_a / "file1.xlsx")

    # ========== folder_a: file2.xlsx ==========
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail_1"
    for col, val in headers.items():
        ws.cell(row=1, column=col, value=val)

    rows_a2 = [
        {2: r"D:\scripts\util.py",     18: "def do_stuff():",    19: "line 5"},
        {2: None,                       18: "import os",          19: "line 1"},
        {2: "/opt/app/run.sh",          18: "#!/bin/bash",        19: "line 1 shebang"},
    ]
    for i, row_data in enumerate(rows_a2, start=2):
        for col, val in row_data.items():
            if val is not None:
                ws.cell(row=i, column=col, value=val)
    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    wb.save(folder_a / "file2.xlsx")

    # ========== folder_b: file1.xlsx (同名) ==========
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail_1"
    for col, val in headers.items():
        ws.cell(row=1, column=col, value=val)

    rows_b1 = [
        # 与 folder_a 完全重复
        {2: r"C:\project\src\main.java", 18: "password=12345", 19: "line 10 context"},
        # 仅文件名相同但内容不同 -> False
        {2: r"/other/path/main.java", 18: "different content", 19: "different context"},
        # 文件名、内容相同但上下文不同 -> False (因为三个字段都参与对比)
        {2: r"C:\project\src\main.java", 18: r"SELECT * FROM users", 19: "line 999 different"},
        # 与 folder_a file1 row4 完全重复
        {2: None, 18: "api_key = \"abcxyz\"", 19: "line 30 // 中文注释"},
        # 全新行 -> False
        {2: "/new/file.cpp", 18: "new content", 19: "new context"},
    ]
    # row2-4 是 C:\project\src\main.java 的合并
    # row5 是新文件，所以 B 列只有 row5 有值
    for i, row_data in enumerate(rows_b1, start=2):
        for col, val in row_data.items():
            if val is not None:
                ws.cell(row=i, column=col, value=val)
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=2)

    wb.save(folder_b / "file1.xlsx")

    # ========== folder_b: file3.xlsx (额外文件, folder_a 中不存在) ==========
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail_1"
    for col, val in headers.items():
        ws.cell(row=1, column=col, value=val)
    rows_b3 = [
        {2: "/opt/app/run.sh", 18: "#!/bin/bash", 19: "line 1 shebang"},       # 重复 (folder_a file2 row3)
        {2: "/opt/app/run.sh", 18: "echo hello", 19: "line 10"},               # 新
        {2: "/unknown/file.txt", 18: "some content", 19: "some context"},      # 新
    ]
    for i, row_data in enumerate(rows_b3, start=2):
        for col, val in row_data.items():
            if val is not None:
                ws.cell(row=i, column=col, value=val)
    wb.save(folder_b / "file3.xlsx")

    print("测试数据已生成:")
    print(f"  folder_a: {folder_a}")
    for f in sorted(folder_a.glob("*.xlsx")):
        print(f"    {f.name}")
    print(f"  folder_b: {folder_b}")
    for f in sorted(folder_b.glob("*.xlsx")):
        print(f"    {f.name}")


if __name__ == "__main__":
    make_test_data()
